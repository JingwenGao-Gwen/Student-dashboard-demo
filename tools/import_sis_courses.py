#!/usr/bin/env python3
"""
Import SIS course catalogue and outline data into MySQL.

Usage:
  python tools/import_sis_courses.py \
    --workbook course_list_database/sis_course_outlines_export.xlsx

Database connection uses MYSQL_URL, or DB_HOST/DB_PORT/DB_NAME/DB_USER/DB_PASSWORD.
"""

from __future__ import annotations

import argparse
import os
import re
from decimal import Decimal, InvalidOperation
from pathlib import Path
from urllib.parse import urlparse

import mysql.connector
from openpyxl import load_workbook


EXPECTED_COLUMNS = [
    "letter",
    "subject",
    "course_code",
    "course_number",
    "course_title",
    "status",
    "units",
    "grading_basis",
    "component",
    "campus",
    "school",
    "academic_org",
    "detail_description",
    "language_of_instruction",
    "description_english",
    "description_chinese",
    "prerequisites",
    "co_requisites",
    "learning_outcomes",
    "course_syllabus",
    "assessment_scheme",
    "grade_type",
    "course_components",
    "terms_status",
    "offered_terms",
    "terms_text",
    "source_folder",
]

ORG_ABBREVIATIONS = {
    "\u7ecf\u7ba1\u5b66\u9662": "SME",
    "\u7406\u5de5\u5b66\u9662": "SSE",
    "\u97f3\u4e50\u5b66\u9662": "MUS",
    "\u6570\u636e\u79d1\u5b66\u5b66\u9662": "SDS",
    "\u533b\u5b66\u9662": "MED",
    "\u4eba\u6587\u793e\u79d1\u5b66\u9662": "HSS",
    "\u516c\u5171\u653f\u7b56\u5b66\u9662": "HSS",
    "\u4eba\u5de5\u667a\u80fd\u5b66\u9662": "SAI",
}

ORG_BY_SUBJECT = {
    "AIE": "SAI",
    "AIR": "SDS", "CSC": "SDS", "DDA": "SDS", "MDS": "SDS", "MFE": "SDS", "RMS": "SDS", "STA": "SDS",
    "ACT": "SME", "DMS": "SME", "ECO": "SME", "FEMBA": "SME", "FEX": "SME", "FIN": "SME",
    "IBA": "SME", "IDE": "SME", "MBM": "SME", "MGT": "SME", "MIS": "SME", "MKT": "SME",
    "BIM": "MED", "BIO": "MED", "BME": "MED", "MED": "MED", "PHM": "MED",
    "MUS": "MUS",
    "CHM": "SSE", "CIE": "SSE", "ECE": "SSE", "EIE": "SSE", "ENE": "SSE", "ERG": "SSE",
    "FMA": "SSE", "FTE": "SSE", "ICS": "SSE", "MAT": "SSE", "MCE": "SSE", "MFM": "SSE",
    "MSE": "SSE", "PHY": "SSE", "SSE": "SSE",
    "CEC": "HSS", "CHI": "HSS", "CLC": "HSS", "CSS": "HSS", "DAI": "HSS", "ENB": "HSS",
    "ENG": "HSS", "ENL": "HSS", "FRN": "HSS", "GEA": "HSS", "GEB": "HSS", "GEC": "HSS",
    "GED": "HSS", "GEW": "HSS", "GFH": "HSS", "GFN": "HSS", "GGE": "HSS", "GLB": "HSS",
    "HSS": "HSS", "ITE": "HSS", "JPN": "HSS", "KOR": "HSS", "LIT": "HSS", "PED": "HSS",
    "PSY": "HSS", "PUB": "HSS", "SPN": "HSS", "SUD": "HSS", "TRA": "HSS", "URB": "HSS",
    "URM": "HSS",
}


CREATE_COURSES_SQL = """
CREATE TABLE IF NOT EXISTS courses (
  course_code varchar(20) COLLATE utf8mb4_unicode_ci NOT NULL,
  course_code_raw varchar(20) COLLATE utf8mb4_unicode_ci DEFAULT NULL,
  subject_code varchar(20) COLLATE utf8mb4_unicode_ci DEFAULT NULL,
  subject_name varchar(255) COLLATE utf8mb4_unicode_ci DEFAULT NULL,
  course_number varchar(20) COLLATE utf8mb4_unicode_ci DEFAULT NULL,
  title varchar(255) COLLATE utf8mb4_unicode_ci DEFAULT NULL,
  title_en varchar(255) COLLATE utf8mb4_unicode_ci DEFAULT NULL,
  title_zh_cn varchar(255) COLLATE utf8mb4_unicode_ci DEFAULT NULL,
  title_zh_tw varchar(255) COLLATE utf8mb4_unicode_ci DEFAULT NULL,
  status varchar(80) COLLATE utf8mb4_unicode_ci DEFAULT NULL,
  units decimal(5,2) DEFAULT NULL,
  grading_basis varchar(120) COLLATE utf8mb4_unicode_ci DEFAULT NULL,
  component varchar(120) COLLATE utf8mb4_unicode_ci DEFAULT NULL,
  campus varchar(120) COLLATE utf8mb4_unicode_ci DEFAULT NULL,
  school varchar(255) COLLATE utf8mb4_unicode_ci DEFAULT NULL,
  academic_org varchar(255) COLLATE utf8mb4_unicode_ci DEFAULT NULL,
  detail_description mediumtext COLLATE utf8mb4_unicode_ci,
  language_of_instruction varchar(120) COLLATE utf8mb4_unicode_ci DEFAULT NULL,
  description_english mediumtext COLLATE utf8mb4_unicode_ci,
  description_chinese mediumtext COLLATE utf8mb4_unicode_ci,
  prerequisites mediumtext COLLATE utf8mb4_unicode_ci,
  co_requisites mediumtext COLLATE utf8mb4_unicode_ci,
  learning_outcomes mediumtext COLLATE utf8mb4_unicode_ci,
  course_syllabus mediumtext COLLATE utf8mb4_unicode_ci,
  assessment_scheme mediumtext COLLATE utf8mb4_unicode_ci,
  grade_type varchar(120) COLLATE utf8mb4_unicode_ci DEFAULT NULL,
  course_components mediumtext COLLATE utf8mb4_unicode_ci,
  terms_status mediumtext COLLATE utf8mb4_unicode_ci,
  offered_terms mediumtext COLLATE utf8mb4_unicode_ci,
  terms_text mediumtext COLLATE utf8mb4_unicode_ci,
  source_folder varchar(255) COLLATE utf8mb4_unicode_ci DEFAULT NULL,
  has_outline tinyint(1) NOT NULL DEFAULT 0,
  source_sheet varchar(40) COLLATE utf8mb4_unicode_ci NOT NULL DEFAULT 'All Courses',
  updated_at timestamp NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
  PRIMARY KEY (course_code),
  KEY idx_courses_subject_code (subject_code),
  KEY idx_courses_school (school),
  KEY idx_courses_has_outline (has_outline)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
"""


UPSERT_SQL = """
INSERT INTO courses (
  course_code, course_code_raw, subject_code, subject_name, course_number,
  title, title_en, title_zh_cn, title_zh_tw,
  status, units, grading_basis, component, campus, school, academic_org,
  detail_description, language_of_instruction, description_english,
  description_chinese, prerequisites, co_requisites, learning_outcomes,
  course_syllabus, assessment_scheme, grade_type, course_components,
  terms_status, offered_terms, terms_text, source_folder, has_outline,
  source_sheet
) VALUES (
  %(course_code)s, %(course_code_raw)s, %(subject_code)s, %(subject_name)s,
  %(course_number)s, %(title)s, %(title_en)s, %(title_zh_cn)s, %(title_zh_tw)s,
  %(status)s, %(units)s, %(grading_basis)s, %(component)s, %(campus)s,
  %(school)s, %(academic_org)s,
  %(detail_description)s, %(language_of_instruction)s, %(description_english)s,
  %(description_chinese)s, %(prerequisites)s, %(co_requisites)s,
  %(learning_outcomes)s, %(course_syllabus)s, %(assessment_scheme)s,
  %(grade_type)s, %(course_components)s, %(terms_status)s, %(offered_terms)s,
  %(terms_text)s, %(source_folder)s, %(has_outline)s, %(source_sheet)s
)
ON DUPLICATE KEY UPDATE
  course_code_raw = VALUES(course_code_raw),
  subject_code = VALUES(subject_code),
  subject_name = VALUES(subject_name),
  course_number = VALUES(course_number),
  title = VALUES(title),
  title_en = VALUES(title_en),
  title_zh_cn = VALUES(title_zh_cn),
  title_zh_tw = VALUES(title_zh_tw),
  status = VALUES(status),
  units = VALUES(units),
  grading_basis = VALUES(grading_basis),
  component = VALUES(component),
  campus = VALUES(campus),
  school = VALUES(school),
  academic_org = VALUES(academic_org),
  detail_description = VALUES(detail_description),
  language_of_instruction = VALUES(language_of_instruction),
  description_english = VALUES(description_english),
  description_chinese = VALUES(description_chinese),
  prerequisites = VALUES(prerequisites),
  co_requisites = VALUES(co_requisites),
  learning_outcomes = VALUES(learning_outcomes),
  course_syllabus = VALUES(course_syllabus),
  assessment_scheme = VALUES(assessment_scheme),
  grade_type = VALUES(grade_type),
  course_components = VALUES(course_components),
  terms_status = VALUES(terms_status),
  offered_terms = VALUES(offered_terms),
  terms_text = VALUES(terms_text),
  source_folder = VALUES(source_folder),
  has_outline = VALUES(has_outline),
  source_sheet = VALUES(source_sheet)
"""

ALTER_COURSES_SQL = [
    "ALTER TABLE courses ADD COLUMN title_en varchar(255) COLLATE utf8mb4_unicode_ci DEFAULT NULL AFTER title",
    "ALTER TABLE courses ADD COLUMN title_zh_cn varchar(255) COLLATE utf8mb4_unicode_ci DEFAULT NULL AFTER title_en",
    "ALTER TABLE courses ADD COLUMN title_zh_tw varchar(255) COLLATE utf8mb4_unicode_ci DEFAULT NULL AFTER title_zh_cn",
    "ALTER TABLE courses MODIFY terms_status mediumtext COLLATE utf8mb4_unicode_ci",
    "ALTER TABLE courses MODIFY offered_terms mediumtext COLLATE utf8mb4_unicode_ci",
]


def db_config() -> dict:
    mysql_url = os.getenv("MYSQL_URL", "").strip()
    if mysql_url:
        parsed = urlparse(mysql_url)
        cfg = {
            "host": parsed.hostname or "localhost",
            "port": int(parsed.port or 3306),
            "database": (parsed.path or "/").lstrip("/") or "course_db",
            "user": parsed.username or "root",
            "password": parsed.password or "",
        }
    else:
        cfg = {
            "host": os.getenv("DB_HOST", "localhost"),
            "port": int(os.getenv("DB_PORT", "3306")),
            "database": os.getenv("DB_NAME", "course_db"),
            "user": os.getenv("DB_USER", "root"),
            "password": os.getenv("DB_PASSWORD", ""),
        }

    if os.getenv("DB_SSL_DISABLED", "false").lower() != "true":
        cfg["ssl_disabled"] = False
    return cfg


def clean(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def display_code(raw: str | None) -> str | None:
    if not raw:
        return None
    code = re.sub(r"\s+", "", str(raw).strip().upper())
    return re.sub(r"^([A-Z]+)(\d)", r"\1 \2", code)


def code_key(value: str | None) -> str:
    return re.sub(r"\s+", "", str(value or "").strip().upper())


def build_title_lookup(workbook: Path) -> dict[str, dict[str, str | None]]:
    """Read a course-list workbook for English/Chinese titles by course code."""
    if not workbook.exists():
        return {}

    wb = load_workbook(workbook, read_only=True, data_only=True)
    lookup: dict[str, dict[str, str | None]] = {}
    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        try:
            headers = [clean(v) or "" for v in next(rows)]
        except StopIteration:
            continue

        lower = [h.lower() for h in headers]
        if "course code" in lower:
            code_idx = lower.index("course code")
            en_idx = lower.index("course title (english)") if "course title (english)" in lower else None
            zh_idx = lower.index("course title (chinese)") if "course title (chinese)" in lower else None
            for row in rows:
                key = code_key(row[code_idx] if code_idx < len(row) else None)
                if not key:
                    continue
                lookup.setdefault(key, {})
                if en_idx is not None and en_idx < len(row):
                    lookup[key]["en"] = lookup[key].get("en") or clean(row[en_idx])
                if zh_idx is not None and zh_idx < len(row):
                    lookup[key]["zh"] = lookup[key].get("zh") or clean(row[zh_idx])
            continue

        # Course_List_byInitial.xlsx uses duplicate course_title headers:
        # subject, course_number, Chinese title, English title, action_id.
        if len(headers) >= 4 and lower[:4] == ["subject", "course_number", "course_title", "course_title"]:
            for row in rows:
                subject = clean(row[0] if len(row) > 0 else None)
                number = clean(row[1] if len(row) > 1 else None)
                if not subject or not number:
                    continue
                subject_code = clean(subject.split(" - ", 1)[0])
                key = code_key(f"{subject_code}{number}")
                if not key:
                    continue
                lookup.setdefault(key, {})
                lookup[key]["zh"] = lookup[key].get("zh") or clean(row[2] if len(row) > 2 else None)
                lookup[key]["en"] = lookup[key].get("en") or clean(row[3] if len(row) > 3 else None)

    return lookup


def subject_parts(subject: str | None) -> tuple[str | None, str | None]:
    text = clean(subject)
    if not text:
        return None, None
    if " - " in text:
        code, name = text.split(" - ", 1)
        return clean(code), clean(name)
    return clean(text.split()[0]), text


def academic_org_abbrev(value: str | None, subject_code: str | None) -> str | None:
    text = clean(value)
    subject = clean(subject_code)
    if text in {"SDS", "SME", "SSE", "MED", "MUS", "HSS", "SAI"}:
        return text
    if text in ORG_ABBREVIATIONS:
        return ORG_ABBREVIATIONS[text]
    if subject in ORG_BY_SUBJECT:
        return ORG_BY_SUBJECT[subject]
    return text


def decimal_or_none(value) -> Decimal | None:
    text = clean(value)
    if text is None:
        return None
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def iter_rows(workbook: Path, sheet_name: str):
    wb = load_workbook(workbook, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Workbook does not contain sheet {sheet_name!r}. Found: {wb.sheetnames}")
    ws = wb[sheet_name]
    headers = [clean(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    missing = [col for col in EXPECTED_COLUMNS if col not in headers]
    if missing:
        raise ValueError(f"Missing expected columns: {missing}")
    for cells in ws.iter_rows(min_row=2):
        row = {headers[i]: clean(cell.value) for i, cell in enumerate(cells) if i < len(headers)}
        if row.get("course_code"):
            yield row


def normalize_row(row: dict, source_sheet: str, title_lookup: dict[str, dict[str, str | None]]) -> dict:
    subject_code, subject_name = subject_parts(row.get("subject"))
    status = row.get("status")
    raw_code = row.get("course_code")
    titles = title_lookup.get(code_key(raw_code), {})
    title_zh = titles.get("zh") or row.get("course_title")
    title_en = titles.get("en")
    return {
        "course_code": display_code(raw_code),
        "course_code_raw": clean(raw_code),
        "subject_code": subject_code,
        "subject_name": subject_name,
        "course_number": row.get("course_number"),
        "title": title_zh,
        "title_en": title_en,
        "title_zh_cn": title_zh,
        "title_zh_tw": title_zh,
        "status": status,
        "units": decimal_or_none(row.get("units")),
        "grading_basis": row.get("grading_basis"),
        "component": row.get("component"),
        "campus": row.get("campus"),
        "school": row.get("school"),
        "academic_org": academic_org_abbrev(row.get("academic_org"), subject_code),
        "detail_description": row.get("detail_description"),
        "language_of_instruction": row.get("language_of_instruction"),
        "description_english": row.get("description_english"),
        "description_chinese": row.get("description_chinese"),
        "prerequisites": row.get("prerequisites"),
        "co_requisites": row.get("co_requisites"),
        "learning_outcomes": row.get("learning_outcomes"),
        "course_syllabus": row.get("course_syllabus"),
        "assessment_scheme": row.get("assessment_scheme"),
        "grade_type": row.get("grade_type"),
        "course_components": row.get("course_components"),
        "terms_status": row.get("terms_status"),
        "offered_terms": row.get("offered_terms"),
        "terms_text": row.get("terms_text"),
        "source_folder": row.get("source_folder"),
        "has_outline": 1 if status == "saved_outline" else 0,
        "source_sheet": source_sheet,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--workbook",
        default="course_list_database/sis_course_outlines_export.xlsx",
        help="Path to sis_course_outlines_export.xlsx",
    )
    parser.add_argument("--sheet", default="All Courses", help="Sheet to import")
    parser.add_argument(
        "--title-workbook",
        default="course_list_database/Course_List_byInitial.xlsx",
        help="Workbook containing English/Chinese course titles keyed by subject + course number.",
    )
    parser.add_argument("--batch-size", type=int, default=500)
    parser.add_argument(
        "--keep-existing",
        action="store_true",
        help="Do not remove courses that are not present in the selected sheet.",
    )
    args = parser.parse_args()

    workbook = Path(args.workbook)
    if not workbook.exists():
        raise FileNotFoundError(workbook)
    title_lookup = build_title_lookup(Path(args.title_workbook))

    conn = mysql.connector.connect(**db_config())
    cursor = conn.cursor()
    cursor.execute(CREATE_COURSES_SQL)
    for statement in ALTER_COURSES_SQL:
        try:
            cursor.execute(statement)
        except mysql.connector.Error as exc:
            if exc.errno != 1060:  # duplicate column
                raise

    rows_by_code = {}
    total = 0
    with_outline = 0
    for raw in iter_rows(workbook, args.sheet):
        row = normalize_row(raw, args.sheet, title_lookup)
        if not row["course_code"]:
            continue
        existing = rows_by_code.get(row["course_code"])
        if existing and existing["has_outline"] and not row["has_outline"]:
            continue
        if not existing or row["has_outline"] or not existing["has_outline"]:
            rows_by_code[row["course_code"]] = row

    batch = []
    for row in rows_by_code.values():
        batch.append(row)
        with_outline += row["has_outline"]
        if len(batch) >= args.batch_size:
            cursor.executemany(UPSERT_SQL, batch)
            total += len(batch)
            conn.commit()
            batch.clear()

    if batch:
        cursor.executemany(UPSERT_SQL, batch)
        total += len(batch)
        conn.commit()

    if not args.keep_existing:
        cursor.execute("SET FOREIGN_KEY_CHECKS = 0")
        cursor.execute("DELETE FROM courses WHERE source_sheet <> %s", (args.sheet,))
        cursor.execute("SET FOREIGN_KEY_CHECKS = 1")
        conn.commit()

    cursor.close()
    conn.close()
    print(f"Imported {total} courses from {workbook} / {args.sheet}.")
    print(f"Courses with saved outline: {with_outline}.")


if __name__ == "__main__":
    main()
