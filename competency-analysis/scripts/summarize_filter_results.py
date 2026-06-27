from __future__ import annotations

import collections
import json
import sys
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
COURSE_XLSX = Path(r"C:\Users\Lenovo\Downloads\ToDesk\sis_course_outlines_export.xlsx")


def main() -> None:
    letters = list((sys.argv[1] if len(sys.argv) > 1 else "IJKLMNOPQRSTUVWXYZ").upper())
    wb = load_workbook(COURSE_XLSX, read_only=True, data_only=True)
    ws = wb["With Outline"]
    headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    idx = {h: i for i, h in enumerate(headers)}
    with_outline = collections.Counter()
    with_syllabus = collections.Counter()
    courses_by_letter: dict[str, dict[str, dict]] = {letter: {} for letter in letters}

    for row in ws.iter_rows(min_row=2, values_only=True):
        letter = str(row[idx["letter"]] or "").upper()
        if letter not in letters:
            continue
        with_outline[letter] += 1
        if str(row[idx["course_syllabus"]] or "").strip():
            with_syllabus[letter] += 1
            code = str(row[idx["course_code"]] or "").strip().upper().replace(" ", "")
            courses_by_letter[letter][code] = {
                "title": row[idx.get("course_title")],
                "school": row[idx.get("school")],
                "subject": row[idx.get("subject")],
            }

    print("Letter\tWithOutline\tWithSyllabus\tScoredCourses\tScoreRows\tExcluded\tExcludedCodes")
    for letter in letters:
        path = ROOT / "outputs" / f"{letter.lower()}_course_faculty_pool_relevance.json"
        data = json.loads(path.read_text(encoding="utf-8"))
        summary = data["summary"]
        excluded = data.get("excluded_ucore_courses", [])
        scored_codes = {row["course_code"] for row in data.get("scored_rows", [])}
        excluded_codes = {item.get("course_code") for item in excluded}
        unscored = sorted(set(courses_by_letter[letter]) - scored_codes - excluded_codes)
        unscored_by_school = collections.Counter(
            courses_by_letter[letter][code].get("school") or ""
            for code in unscored
        )
        codes = [
            f"{item.get('course_code')} ({item.get('exclude_reason', '')})"
            for item in excluded
        ]
        code_text = "; ".join(codes[:30])
        if len(codes) > 30:
            code_text += f"; ... (+{len(codes) - 30} more)"
        print(
            f"{letter}\t{with_outline[letter]}\t{with_syllabus[letter]}\t"
            f"{summary['courses_scored']}\t{summary['score_rows']}\t"
            f"{summary['ucore_courses_excluded']}\t{code_text}"
        )
        if unscored:
            school_text = "; ".join(f"{school or '[blank]'}={count}" for school, count in unscored_by_school.items())
            sample_text = ", ".join(unscored[:20])
            if len(unscored) > 20:
                sample_text += f", ... (+{len(unscored) - 20} more)"
            print(f"{letter}\tUNSCORED_NONEXCLUDED\t{len(unscored)}\t{school_text}\t{sample_text}")


if __name__ == "__main__":
    main()
