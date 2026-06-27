from __future__ import annotations

import json
import math
import re
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader


ROOT = Path(__file__).resolve().parents[1]
COURSE_XLSX = Path(r"C:\Users\Lenovo\Downloads\ToDesk\sis_course_outlines_export.xlsx")
COMPETENCY_XLSX = ROOT / "outputs" / "major_competency_ai_review.xlsx"
CURRICULUM_PDFS = [
    Path(r"C:\Users\Lenovo\Downloads\Academic Curriculum of the Undergraduate Programmes (for students admitted in 2025-26 and thereafter)_Circular(AB202.pdf"),
    Path(r"C:\Users\Lenovo\Downloads\Academic Curriculum of the Undergraduate Programmes (for students admitted in 2023-24 to 2024-25)_3rd(2025)_20250714.pdf"),
]
OUT_JSON = ROOT / "outputs" / "a_course_competency_relevance.json"


SCHOOL_MAP = {
    "经管学院": "School of Management and Economics",
    "数据科学学院": "School of Data Science",
    "人工智能学院": "School of Artificial Intelligence",
    "理工学院": "School of Science and Engineering",
    "医学院": "School of Medicine",
    "人文社科学院": "School of Humanities and Social Science",
    "音乐学院": "School of Music",
}

SCHOOL_ABBR = {
    "School of Management and Economics": "SME",
    "School of Data Science": "SDS",
    "School of Artificial Intelligence": "SAI",
    "School of Science and Engineering": "SSE",
    "School of Medicine": "MED",
    "School of Humanities and Social Science": "HSS",
    "School of Music": "MUS",
}

STOPWORDS = {
    "and", "the", "for", "with", "this", "that", "from", "into", "will", "are", "can", "may",
    "their", "such", "through", "course", "courses", "student", "students", "major", "programme",
    "program", "foundation", "foundations", "basic", "core", "required", "elective", "electives",
    "ability", "abilities", "competency", "competencies", "knowledge", "skills", "skill",
    "understanding", "including", "using", "used", "use", "based", "related", "advanced",
    "introduction", "introductory", "principles", "concepts", "methods", "applications",
    "analysis", "analytical", "professional", "decision", "making",
}

DOMAIN_TERMS = {
    "accounting": ["accounting", "accountancy", "financial statement", "financial reporting", "audit", "tax", "cost", "management accounting", "corporate governance", "valuation", "forensic", "fraud"],
    "finance": ["finance", "financial", "capital market", "investment", "valuation", "risk", "credit", "derivative", "asset pricing", "corporate finance", "financial market"],
    "economics": ["economics", "economic", "microeconomics", "macroeconomics", "market", "policy", "econometric"],
    "business": ["business", "management", "marketing", "strategy", "organization", "operation", "managerial", "corporate", "supply chain"],
    "data": ["data", "analytics", "mining", "database", "visualization", "machine learning", "statistical", "statistics", "programming", "algorithm", "text analysis"],
    "ai": ["artificial intelligence", "ai", "machine learning", "deep learning", "reinforcement learning", "large language model", "optimization", "robotics", "neural"],
    "computing": ["programming", "software", "algorithm", "computer", "systems", "database", "data structures", "nlp"],
    "research_project": ["research", "project", "capstone", "internship", "practice", "seminar"],
}

PROGRAMME_FOCUS = {
    "Professional Accountancy": ["accounting", "finance", "business", "data", "research_project"],
    "Finance": ["finance", "accounting", "economics", "data"],
    "Financial Engineering": ["finance", "data", "computing", "economics"],
    "Economics": ["economics", "finance", "data", "business"],
    "Global Business Studies": ["business", "finance", "economics"],
    "Big Data Management and Applications": ["data", "business", "computing", "finance"],
    "Marketing and Communication": ["business", "data"],
    "Artificial Intelligence": ["ai", "computing", "data", "research_project"],
    "Data Science and Big Data Technology": ["data", "computing", "ai", "research_project"],
    "Computer Science and Engineering": ["computing", "ai", "data", "research_project"],
    "Statistics": ["data", "research_project"],
    "X & Interdisciplinary Data Analytics Double Major": ["data", "business", "finance", "economics", "computing"],
}

PREFIX_BOOST = {
    "ACT": {
        "Professional Accountancy": 2.8,
        "Finance": 1.5,
        "Financial Engineering": 1.2,
        "Economics": 0.8,
        "Global Business Studies": 1.0,
        "Big Data Management and Applications": 1.0,
        "Marketing and Communication": 0.4,
        "X & Interdisciplinary Data Analytics Double Major": 0.7,
    },
    "AIE": {"Artificial Intelligence": 3.0},
    "AIR": {
        "Artificial Intelligence": 2.2,
        "Data Science and Big Data Technology": 1.7,
        "Computer Science and Engineering": 1.7,
        "Statistics": 0.8,
        "Financial Engineering": 0.6,
        "X & Interdisciplinary Data Analytics Double Major": 1.4,
    },
}


def normalize_code(code: str) -> str:
    return re.sub(r"\s+", "", code or "").upper()


def extract_ucore_codes() -> dict:
    by_pdf = {}
    all_codes = set()
    pattern = re.compile(r"\b[A-Z]{2,4}\s?\d{4}[A-Z]?\b")
    for pdf in CURRICULUM_PDFS:
        reader = PdfReader(str(pdf))
        text = "\n".join((page.extract_text() or "") for page in reader.pages)
        codes = {normalize_code(c) for c in pattern.findall(text)}
        by_pdf[pdf.name] = sorted(codes)
        all_codes.update(codes)
    return {"all_codes": sorted(all_codes), "by_pdf": by_pdf}


def tokenize(text: str) -> list[str]:
    text = (text or "").lower()
    text = text.replace("financial statements", "financial_statement")
    text = text.replace("financial reporting", "financial_reporting")
    text = text.replace("capital market", "capital_market")
    raw = re.findall(r"[a-z][a-z_]{2,}", text)
    return [t for t in raw if t not in STOPWORDS and len(t) > 2]


def phrase_hits(text: str, programme: str) -> list[str]:
    lower = (text or "").lower()
    areas = PROGRAMME_FOCUS.get(programme, [])
    hits = []
    for area in areas:
        for phrase in DOMAIN_TERMS.get(area, []):
            if phrase in lower:
                hits.append(phrase)
    return sorted(set(hits))


def read_courses(ucore_codes: set[str]) -> tuple[list[dict], list[dict]]:
    wb = load_workbook(COURSE_XLSX, read_only=True, data_only=True)
    ws = wb["With Outline"]
    headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    idx = {h: i for i, h in enumerate(headers)}
    courses = []
    excluded = []
    text_fields = [
        "course_title", "detail_description", "description_english", "description_chinese",
        "prerequisites", "co_requisites", "learning_outcomes", "course_syllabus",
        "assessment_scheme", "course_components", "offered_terms",
    ]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[idx["letter"]] != "A":
            continue
        code = normalize_code(row[idx["course_code"]])
        record = {h: row[i] for h, i in idx.items()}
        record["course_code"] = code
        record["school_en"] = SCHOOL_MAP.get(record.get("school") or "", record.get("school") or "")
        record["text_for_scoring"] = "\n".join(str(record.get(f) or "") for f in text_fields)
        if code in ucore_codes:
            record["exclude_reason"] = "UCore course listed in supplied Academic Curriculum PDF(s)"
            excluded.append(record)
        else:
            courses.append(record)
    return courses, excluded


def read_competencies() -> list[dict]:
    wb = load_workbook(COMPETENCY_XLSX, read_only=True, data_only=True)
    ws = wb["Competency Review"]
    headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    idx = {h: i for i, h in enumerate(headers)}
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[idx["Programme"]]:
            continue
        rec = {h: row[i] if i < len(row) else None for h, i in idx.items()}
        text = "\n".join(str(rec.get(h) or "") for h in [
            "Basic Professional Foundation",
            "Core Competencies Expected",
            "Elective / Stream-Based Competencies",
            "Curriculum Evidence",
        ])
        rec["competency_text"] = text
        records.append(rec)
    return records


def programme_applies_to_school(programme: dict, school_en: str) -> bool:
    if programme.get("School / Faculty") == school_en:
        return True
    note = f"{programme.get('School Notes') or ''} {programme.get('School / Faculty') or ''}"
    abbr = SCHOOL_ABBR.get(school_en)
    if abbr and re.search(rf"\b{re.escape(abbr)}\b", note):
        return True
    if programme.get("Programme") == "Financial Engineering" and school_en in {
        "School of Management and Economics", "School of Data Science", "School of Science and Engineering"
    }:
        return True
    return False


def build_idf(programmes: list[dict]) -> dict[str, float]:
    docs = [set(tokenize(p["competency_text"])) for p in programmes]
    df = Counter()
    for doc in docs:
        df.update(doc)
    n = max(len(docs), 1)
    return {term: math.log((n + 1) / (freq + 1)) + 1 for term, freq in df.items()}


def score_pair(course: dict, programme: dict, idf: dict[str, float]) -> dict:
    course_text = course["text_for_scoring"]
    programme_text = programme["competency_text"]
    course_tokens = Counter(tokenize(course_text))
    prog_tokens = Counter(tokenize(programme_text))
    ranked_prog = sorted(prog_tokens, key=lambda t: prog_tokens[t] * idf.get(t, 1.0), reverse=True)[:70]
    denom = sum(prog_tokens[t] * idf.get(t, 1.0) for t in ranked_prog) or 1.0
    matched = [t for t in ranked_prog if t in course_tokens]
    overlap = sum(min(course_tokens[t], prog_tokens[t]) * idf.get(t, 1.0) for t in matched) / denom

    phrase_match = phrase_hits(course_text, programme["Programme"])
    phrase_component = min(2.0, len(phrase_match) * 0.28)
    prefix = re.match(r"[A-Z]+", course["course_code"]).group(0)
    prefix_component = PREFIX_BOOST.get(prefix, {}).get(programme["Programme"], 0.0)
    text_component = min(6.2, 6.2 * math.sqrt(max(overlap, 0)))

    # Graduate-level courses can still support undergraduate competencies, but they should not
    # automatically dominate the fit score only because they share a prefix.
    level = int(re.search(r"\d", course["course_code"]).group(0)) if re.search(r"\d", course["course_code"]) else 0
    grad_penalty = 0.4 if level >= 5 and phrase_component < 0.7 else 0.0
    score = max(0.0, min(10.0, text_component + phrase_component + prefix_component - grad_penalty))

    if score >= 8:
        level_label = "Very high"
    elif score >= 6:
        level_label = "High"
    elif score >= 4:
        level_label = "Moderate"
    elif score >= 2:
        level_label = "Low"
    else:
        level_label = "Very low"

    top_terms = sorted(set(matched[:10] + [p.replace(" ", "_") for p in phrase_match[:8]]))
    rationale_bits = []
    if prefix_component:
        rationale_bits.append(f"{prefix} prefix aligns with {programme['Programme']} ({prefix_component:.1f} boost)")
    if phrase_match:
        rationale_bits.append("course outline mentions " + ", ".join(phrase_match[:6]))
    if matched:
        rationale_bits.append("shared competency terms include " + ", ".join(matched[:6]))
    if not rationale_bits:
        rationale_bits.append("limited explicit match with this programme competency profile")

    return {
        "score": round(score, 1),
        "relevance_level": level_label,
        "matched_terms": ", ".join(top_terms),
        "rationale": "; ".join(rationale_bits),
        "text_similarity_component": round(text_component, 2),
        "phrase_component": round(phrase_component, 2),
        "prefix_component": round(prefix_component, 2),
    }


def main() -> None:
    ucore = extract_ucore_codes()
    courses, excluded = read_courses(set(ucore["all_codes"]))
    programmes = read_competencies()
    idf = build_idf(programmes)
    rows = []
    for course in courses:
        applicable = [p for p in programmes if programme_applies_to_school(p, course["school_en"])]
        for programme in applicable:
            scoring = score_pair(course, programme, idf)
            rows.append({
                "course_code": course["course_code"],
                "course_title": course.get("course_title") or "",
                "subject": course.get("subject") or "",
                "offering_school": course.get("school") or "",
                "offering_school_en": course.get("school_en") or "",
                "academic_org": course.get("academic_org") or "",
                "programme": programme["Programme"],
                "programme_school": programme.get("School / Faculty") or "",
                "programme_school_notes": programme.get("School Notes") or "",
                **scoring,
                "course_description": course.get("description_english") or course.get("detail_description") or "",
                "learning_outcomes": course.get("learning_outcomes") or "",
                "course_syllabus": course.get("course_syllabus") or "",
            })
    rows.sort(key=lambda r: (r["course_code"], -r["score"], r["programme"]))
    OUT_JSON.write_text(json.dumps({
        "method": {
            "scope": "A-letter SIS courses with saved course outlines; UCore courses appearing in supplied Academic Curriculum PDFs are excluded from scoring.",
            "score_scale": "0-10",
            "score_components": ["outline/competency text similarity", "domain phrase matches", "course prefix/programme alignment"],
        },
        "ucore_codes": ucore,
        "scored_rows": rows,
        "excluded_ucore_courses": excluded,
        "summary": {
            "courses_scored": len({r["course_code"] for r in rows}),
            "score_rows": len(rows),
            "ucore_courses_excluded": len(excluded),
        },
    }, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"out": str(OUT_JSON), "courses": len({r["course_code"] for r in rows}), "rows": len(rows), "excluded": len(excluded)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
