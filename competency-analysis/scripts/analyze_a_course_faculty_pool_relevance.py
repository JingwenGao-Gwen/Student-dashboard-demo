from __future__ import annotations

import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader


ROOT = Path(__file__).resolve().parents[1]
COURSE_XLSX = Path(r"C:\Users\Lenovo\Downloads\ToDesk\sis_course_outlines_export.xlsx")
COMPETENCY_XLSX = ROOT / "outputs" / "major_competency_ai_review.xlsx"
CURRICULUM_PDFS = [
    Path(r"C:\Users\Lenovo\Downloads\Academic Curriculum of the Undergraduate Programmes (for students admitted in 2025-26 and thereafter)_Circular(AB202.pdf"),
    Path(r"C:\Users\Lenovo\Downloads\Academic Curriculum of the Undergraduate Programmes (for students admitted in 2023-24 to 2024-25)_3rd(2025)_20250714.pdf"),
    Path(r"C:\Users\Lenovo\Downloads\香港中文大学（深圳）国情教育课程修读规则_2021-22年度及以后入学学生本科生适用_2024年8月修订.pdf"),
]
TARGET_LETTER = (sys.argv[1] if len(sys.argv) > 1 else "A").strip().upper()
OUT_JSON = ROOT / "outputs" / f"{TARGET_LETTER.lower()}_course_faculty_pool_relevance.json"
PREFIX_EXCLUSION_RULES = {
    "CEC": "CEC course excluded by user rule; not scored for major competency contribution",
    "GEB": "GEB UCore course excluded by user rule; not scored for major competency contribution",
    "GEC": "GEC UCore course excluded by user rule; not scored for major competency contribution",
    "GED": "GED UCore course excluded by user rule; not scored for major competency contribution",
    "GEW": "GEW UCore course excluded by user rule; not scored for major competency contribution",
}


SCHOOL_MAP = {
    "经管学院": "School of Management and Economics",
    "数据科学学院": "School of Data Science",
    "人工智能学院": "School of Artificial Intelligence",
    "理工学院": "School of Science and Engineering",
    "医学院": "School of Medicine",
    "人文社科学院": "School of Humanities and Social Science",
    "音乐学院": "School of Music",
    "公共政策学院": "School of Public Policy",
}

SCHOOL_ABBR = {
    "School of Management and Economics": "SME",
    "School of Data Science": "SDS",
    "School of Artificial Intelligence": "SAI",
    "School of Science and Engineering": "SSE",
    "School of Medicine": "MED",
    "School of Humanities and Social Science": "HSS",
    "School of Music": "MUS",
    "School of Public Policy": "SPP",
}


POOL_DEFINITIONS = {
    "School of Management and Economics": [
        {
            "competency": "Accounting and financial reporting",
            "definition": "Ability to understand accounting information, financial statements, reporting rules, and accounting theory.",
            "keywords": ["accounting", "accountancy", "financial statement", "financial statements", "financial reporting", "reporting", "ledger", "revenue", "liability", "asset", "equity", "ifrs", "gaap"],
            "prefix_boost": {"ACT": 2.2},
        },
        {
            "competency": "Management accounting, cost control and managerial decision-making",
            "definition": "Ability to use cost, budget, performance and internal accounting information for management decisions.",
            "keywords": ["management accounting", "managerial", "cost", "budget", "variance", "performance", "planning", "control", "decision", "internal", "operation", "operations"],
            "prefix_boost": {"ACT": 1.6},
        },
        {
            "competency": "Auditing, assurance, tax, law, ethics and governance",
            "definition": "Ability to reason about assurance, taxation, law, compliance, fraud, business ethics, corporate governance and social responsibility.",
            "keywords": ["audit", "auditing", "assurance", "tax", "taxation", "law", "legal", "compliance", "fraud", "ethics", "ethical", "governance", "corporate governance", "social responsibility", "csr"],
            "prefix_boost": {"ACT": 1.4},
        },
        {
            "competency": "Corporate finance, investment, valuation and risk",
            "definition": "Ability to analyze corporate finance, capital markets, investment, valuation, restructuring, credit and financial risk.",
            "keywords": ["finance", "financial", "capital market", "capital markets", "investment", "valuation", "risk", "credit", "banking", "derivative", "asset pricing", "corporate finance", "restructuring", "merger", "acquisition"],
            "prefix_boost": {"ACT": 0.9},
        },
        {
            "competency": "Economics, markets and econometrics",
            "definition": "Ability to use economic theory, market reasoning, policy logic and econometric/statistical evidence.",
            "keywords": ["economics", "economic", "microeconomics", "macroeconomics", "market", "markets", "policy", "econometric", "econometrics", "demand", "supply", "consumer", "producer"],
            "prefix_boost": {"ACT": 0.2},
        },
        {
            "competency": "Marketing, consumer insight and communication",
            "definition": "Ability to analyze markets, consumers, brands, marketing strategy and communication.",
            "keywords": ["marketing", "consumer", "customer", "brand", "communication", "advertising", "market research", "digital marketing", "sales", "promotion"],
            "prefix_boost": {"ACT": 0.1},
        },
        {
            "competency": "Business management, strategy, operations and supply chain",
            "definition": "Ability to understand organizations, management, business strategy, operations, logistics and supply chains.",
            "keywords": ["business", "management", "organization", "organisational", "organizational", "strategy", "strategic", "operation", "operations", "supply chain", "logistics", "entrepreneurship", "corporate"],
            "prefix_boost": {"ACT": 0.5},
        },
        {
            "competency": "Business data analytics, MIS, AI/data applications and visualization",
            "definition": "Ability to apply data analytics, databases, information systems, AI tools and visualization in business contexts.",
            "keywords": ["data", "analytics", "analysis", "mining", "database", "visualization", "visualisation", "dashboard", "machine learning", "artificial intelligence", "ai", "mis", "information system", "text analysis", "python", "algorithm"],
            "prefix_boost": {"ACT": 0.7},
        },
        {
            "competency": "Quantitative and statistical business reasoning",
            "definition": "Ability to use mathematics, statistics, quantitative methods and modelling for business decisions.",
            "keywords": ["statistics", "statistical", "quantitative", "mathematics", "math", "model", "modelling", "modeling", "probability", "regression", "forecast", "forecasting"],
            "prefix_boost": {"ACT": 0.2},
        },
        {
            "competency": "Research, project execution and workplace readiness",
            "definition": "Ability to conduct research, complete projects, communicate professionally, and connect classroom knowledge to workplace practice.",
            "keywords": ["research", "project", "capstone", "internship", "practice", "seminar", "case", "presentation", "communication", "professional", "workplace", "career"],
            "prefix_boost": {"ACT": 0.1},
        },
    ],
    "School of Artificial Intelligence": [
        {
            "competency": "AI programming and software implementation",
            "definition": "Ability to implement AI solutions through programming, software tools, algorithms and data structures.",
            "keywords": ["programming", "python", "software", "implementation", "algorithm", "data structure", "data structures", "computer", "coding", "program"],
            "prefix_boost": {"AIE": 2.0},
        },
        {
            "competency": "Mathematical and statistical foundations for AI",
            "definition": "Ability to use mathematics, statistics, probability, optimization and modelling for AI.",
            "keywords": ["mathematics", "math", "statistics", "statistical", "probability", "linear algebra", "calculus", "optimization", "optimisation", "model", "modelling"],
            "prefix_boost": {"AIE": 0.7},
        },
        {
            "competency": "Machine learning, deep learning and large language models",
            "definition": "Ability to understand and apply machine learning, deep learning, neural models and large language models.",
            "keywords": ["machine learning", "deep learning", "neural", "large language model", "llm", "language model", "reinforcement learning", "supervised", "unsupervised", "training"],
            "prefix_boost": {"AIE": 1.4},
        },
        {
            "competency": "AI systems, algorithms and data-driven problem solving",
            "definition": "Ability to design AI systems, use algorithms and data to solve applied problems.",
            "keywords": ["ai system", "systems", "algorithm", "data", "problem", "solution", "model", "prediction", "classification", "decision"],
            "prefix_boost": {"AIE": 1.2},
        },
        {
            "competency": "AI ethics, governance and regulation",
            "definition": "Ability to reason about ethics, governance, regulation, safety and social impact of AI.",
            "keywords": ["ethics", "ethical", "governance", "regulation", "policy", "privacy", "fairness", "bias", "safety", "social impact"],
            "prefix_boost": {"AIE": 0.4},
        },
        {
            "competency": "Project-based AI exploration, practice and capstone execution",
            "definition": "Ability to carry out staged AI projects, capstone work, team practice and applied exploration.",
            "keywords": ["project", "capstone", "practice", "exploration", "team", "research", "design", "implementation", "presentation"],
            "prefix_boost": {"AIE": 1.1},
        },
        {
            "competency": "Embodied, robotic and domain AI applications",
            "definition": "Ability to apply AI to robotics, embodied systems, science, business or other domains.",
            "keywords": ["robot", "robotics", "embodied", "control", "sensor", "domain", "science", "business", "application", "automation"],
            "prefix_boost": {"AIE": 0.7},
        },
    ],
    "School of Data Science": [
        {
            "competency": "Programming, algorithms, software and computing systems",
            "definition": "Ability to program, reason about algorithms, build software and understand computing systems.",
            "keywords": ["programming", "python", "software", "algorithm", "algorithms", "computer", "computing", "system", "systems", "implementation", "coding"],
            "prefix_boost": {"AIR": 1.4},
        },
        {
            "competency": "Data structures, databases and computing infrastructure",
            "definition": "Ability to organize, store, retrieve and manage data through data structures, databases and infrastructure.",
            "keywords": ["data structure", "data structures", "database", "databases", "data management", "storage", "query", "infrastructure", "information"],
            "prefix_boost": {"AIR": 0.5},
        },
        {
            "competency": "Statistical reasoning, probability and inference",
            "definition": "Ability to use probability, statistics, inference and uncertainty reasoning.",
            "keywords": ["statistics", "statistical", "probability", "inference", "regression", "estimation", "hypothesis", "uncertainty", "sampling"],
            "prefix_boost": {"AIR": 0.2},
        },
        {
            "competency": "Machine learning, AI and data mining",
            "definition": "Ability to use machine learning, AI, data mining and predictive modelling.",
            "keywords": ["machine learning", "deep learning", "reinforcement learning", "artificial intelligence", "ai", "data mining", "prediction", "classification", "neural", "model", "training"],
            "prefix_boost": {"AIR": 1.5},
        },
        {
            "competency": "Big data analytics and data-driven modelling",
            "definition": "Ability to analyze large-scale data and build data-driven models for decisions.",
            "keywords": ["big data", "data", "analytics", "analysis", "modelling", "modeling", "data-driven", "visualization", "forecast", "decision"],
            "prefix_boost": {"AIR": 0.6},
        },
        {
            "competency": "Quantitative modelling and mathematical foundations",
            "definition": "Ability to use mathematics, optimization, numerical methods and quantitative models.",
            "keywords": ["mathematics", "math", "quantitative", "optimization", "optimisation", "linear algebra", "calculus", "model", "modelling", "control"],
            "prefix_boost": {"AIR": 0.7},
        },
        {
            "competency": "Robotics, control and embodied intelligent systems",
            "definition": "Ability to understand robotics, control, sensors, embodied intelligence and intelligent systems.",
            "keywords": ["robot", "robotics", "control", "sensor", "tactile", "soft robotics", "micro", "nano", "nanotechnology", "embodied", "manipulation"],
            "prefix_boost": {"AIR": 2.0},
        },
        {
            "competency": "Research, capstone and project execution",
            "definition": "Ability to complete research, capstone, project, design and applied team work.",
            "keywords": ["research", "project", "capstone", "design", "team", "practice", "implementation", "presentation", "seminar"],
            "prefix_boost": {"AIR": 0.8},
        },
        {
            "competency": "Domain applications in business, finance, life science and engineering",
            "definition": "Ability to apply data science and AI to business, finance, life science, engineering or other domains.",
            "keywords": ["business", "finance", "financial", "life science", "biology", "biomedical", "engineering", "application", "domain", "materials"],
            "prefix_boost": {"AIR": 0.4},
        },
    ],
    "School of Medicine": [
        {
            "competency": "Biomedical and life science foundation",
            "definition": "Ability to understand core biology, chemistry, life science, human structure/function and biomedical foundations.",
            "keywords": ["biology", "biological", "life science", "biomedical", "chemistry", "biochemistry", "human", "anatomy", "physiology", "cell", "molecular", "organism"],
            "prefix_boost": {"BIO": 1.6, "BIM": 0.7, "BME": 0.9},
        },
        {
            "competency": "Molecular biology, genetics, genomics and proteomics",
            "definition": "Ability to reason about genes, genomes, proteins, molecular mechanisms and omics technologies.",
            "keywords": ["molecular", "genetic", "genetics", "genome", "genomics", "proteomics", "protein", "dna", "rna", "gene", "epigenetics", "sequencing"],
            "prefix_boost": {"BIO": 1.2, "BIM": 1.4},
        },
        {
            "competency": "Computational biology and bioinformatics",
            "definition": "Ability to use computational, algorithmic and database methods for biological and health-science data.",
            "keywords": ["bioinformatics", "computational biology", "computational", "algorithm", "algorithms", "database", "biological database", "sequence", "alignment", "genomic", "protein structure"],
            "prefix_boost": {"BIM": 2.2, "BIO": 0.4, "BME": 0.2},
        },
        {
            "competency": "Biomedical data analytics, statistics and machine learning",
            "definition": "Ability to apply statistics, visualization, machine learning and data analysis to biomedical or biological problems.",
            "keywords": ["statistics", "statistical", "biostatistics", "data", "analytics", "analysis", "visualization", "machine learning", "model", "modelling", "classification", "prediction"],
            "prefix_boost": {"BIM": 1.5, "BIO": 0.6, "BME": 0.8},
        },
        {
            "competency": "Laboratory, experimental and research methods",
            "definition": "Ability to perform biological, biomedical, laboratory and research work, including experimental design and interpretation.",
            "keywords": ["laboratory", "lab", "experiment", "experimental", "hands-on", "research", "method", "methods", "design", "measurement", "sample", "assay"],
            "prefix_boost": {"BIO": 1.2, "BME": 1.0, "BIM": 0.5},
        },
        {
            "competency": "Biomedical engineering systems and instrumentation",
            "definition": "Ability to understand biomedical engineering systems, instrumentation, biosensors, signals, imaging and medical devices.",
            "keywords": ["biomedical engineering", "engineering", "instrumentation", "instrument", "biosensor", "sensor", "medical device", "device", "signal", "imaging", "control", "system"],
            "prefix_boost": {"BME": 2.3, "BIO": 0.1, "BIM": 0.2},
        },
        {
            "competency": "Biomaterials, tissue engineering and regenerative medicine",
            "definition": "Ability to understand biomaterials, tissue engineering, biofabrication, nanobiotechnology and regenerative medicine.",
            "keywords": ["biomaterial", "biomaterials", "tissue engineering", "tissue", "regenerative", "regeneration", "biofabrication", "nanobiotechnology", "nanomaterial", "carrier", "scaffold", "3d"],
            "prefix_boost": {"BME": 1.9, "BIO": 0.3},
        },
        {
            "competency": "Biomedical systems, biomechanics and physiology",
            "definition": "Ability to analyze biological systems, biomechanics, physiology, neural systems and systems-level biomedical phenomena.",
            "keywords": ["physiology", "biomechanics", "biomechanical", "neural", "nervous", "system", "systems", "control", "animal", "organ", "function"],
            "prefix_boost": {"BME": 1.3, "BIO": 0.9},
        },
        {
            "competency": "Clinical, medical, ethical and translational readiness",
            "definition": "Ability to connect biomedical knowledge with clinical practice, medical translation, ethics, internship and professional readiness.",
            "keywords": ["clinical", "medical", "medicine", "ethics", "bioethics", "translation", "translational", "internship", "practice", "hospital", "patient", "diagnosis"],
            "prefix_boost": {"BME": 0.9, "BIO": 0.7},
        },
        {
            "competency": "Capstone, project execution and innovation",
            "definition": "Ability to complete capstone projects, internships, research projects and innovation-oriented biomedical work.",
            "keywords": ["capstone", "project", "internship", "research", "design", "innovation", "practice", "proposal", "presentation", "team"],
            "prefix_boost": {"BME": 0.8, "BIO": 0.8, "BIM": 0.5},
        },
    ],
    "School of Humanities and Social Science": [
        {
            "competency": "Language proficiency, writing and textual communication",
            "definition": "Ability to use language effectively for reading, writing, speaking, analysis and professional communication.",
            "keywords": ["language", "writing", "reading", "speaking", "communication", "text", "textual", "grammar", "vocabulary", "composition", "rhetoric", "presentation"],
            "prefix_boost": {"CHI": 1.4, "CLC": 1.7, "CEC": 0.2, "CSS": 0.2},
        },
        {
            "competency": "Literary, cultural and historical analysis",
            "definition": "Ability to analyze literature, culture, history, classics, arts and cultural texts.",
            "keywords": ["literature", "literary", "culture", "cultural", "history", "classical", "poetry", "fiction", "drama", "art", "film", "heritage", "relics"],
            "prefix_boost": {"CHI": 1.8, "CLC": 1.2},
        },
        {
            "competency": "Chinese language, linguistics and second-language pedagogy",
            "definition": "Ability to understand Chinese linguistics, dialects, phonology, syntax and Chinese language teaching.",
            "keywords": ["chinese", "linguistics", "dialect", "dialects", "phonology", "syntax", "language teaching", "second language", "mandarin", "grammar", "prosodic"],
            "prefix_boost": {"CHI": 1.7, "CLC": 1.8},
        },
        {
            "competency": "Civic, ethical, political and social awareness",
            "definition": "Ability to understand civic education, ethics, political theory, society, institutions and contemporary China.",
            "keywords": ["civic", "ethics", "ethical", "political", "politics", "marxism", "socialism", "society", "social", "citizen", "china", "governance", "ideology"],
            "prefix_boost": {"CEC": 2.2, "CHI": 0.2, "CSS": 0.5},
        },
        {
            "competency": "Social science methodology and computational social science",
            "definition": "Ability to use social science methods, computational tools, statistics and data for social analysis.",
            "keywords": ["social science", "methodology", "method", "methods", "computational", "data", "statistics", "statistical", "model", "survey", "network", "causal"],
            "prefix_boost": {"CSS": 2.2, "CEC": 0.6},
        },
        {
            "competency": "Intercultural, translation and cultural mediation readiness",
            "definition": "Ability to work across cultures, languages and social contexts, including translation or cultural mediation.",
            "keywords": ["intercultural", "translation", "translate", "culture", "cultural", "bilingual", "multilingual", "communication", "mediation", "international"],
            "prefix_boost": {"CHI": 0.8, "CLC": 1.0},
        },
        {
            "competency": "Research, critical inquiry and academic writing",
            "definition": "Ability to conduct humanities/social-science research, develop critical arguments and write academically.",
            "keywords": ["research", "critical", "inquiry", "academic", "paper", "thesis", "method", "methods", "writing", "seminar", "analysis"],
            "prefix_boost": {"CHI": 0.8, "CSS": 0.9, "CEC": 0.5},
        },
    ],
    "School of Music": [
        {
            "competency": "Music theory, musicianship and aural skills",
            "definition": "Ability to understand music theory, harmony, counterpoint, form, notation, sight-singing, ear training and core musicianship.",
            "keywords": ["music theory", "theory", "musicianship", "aural", "ear training", "sight singing", "sight-singing", "harmony", "counterpoint", "form", "notation", "solfege"],
            "prefix_boost": {"MUS": 1.8},
        },
        {
            "competency": "Composition, arrangement and creative production",
            "definition": "Ability to compose, arrange, orchestrate and develop original creative musical work.",
            "keywords": ["composition", "compose", "composing", "composer", "creative", "creativity", "arrangement", "arranging", "orchestration", "songwriting", "original", "production"],
            "prefix_boost": {"MUS": 1.7},
        },
        {
            "competency": "Music analysis, history and contextual understanding",
            "definition": "Ability to analyze musical works and understand music history, styles, genres, cultural context and repertoire.",
            "keywords": ["analysis", "analyze", "analytical", "history", "historical", "style", "genre", "context", "repertoire", "musicology", "ethnomusicology", "culture", "cultural"],
            "prefix_boost": {"MUS": 1.4},
        },
        {
            "competency": "Performance, ensemble and studio practice",
            "definition": "Ability to perform vocally or instrumentally, rehearse, participate in ensemble/studio work and develop practical performance technique.",
            "keywords": ["performance", "perform", "performing", "instrument", "instrumental", "vocal", "voice", "singing", "ensemble", "studio", "rehearsal", "recital", "technique"],
            "prefix_boost": {"MUS": 1.8},
        },
        {
            "competency": "Repertoire interpretation and critical listening",
            "definition": "Ability to interpret repertoire, listen critically, evaluate musical expression and connect performance with stylistic understanding.",
            "keywords": ["repertoire", "interpretation", "interpret", "listening", "critical listening", "expressive", "expression", "style", "phrasing", "artist", "artistic", "aesthetic"],
            "prefix_boost": {"MUS": 1.4},
        },
        {
            "competency": "Music technology, recording and digital production",
            "definition": "Ability to use music technology, recording, audio, digital tools and production methods in music work.",
            "keywords": ["technology", "recording", "audio", "digital", "sound", "production", "studio", "software", "midi", "electronic", "acoustic", "mixing"],
            "prefix_boost": {"MUS": 1.2},
        },
        {
            "competency": "Research, writing and professional artistic practice",
            "definition": "Ability to conduct music research, write about music, present artistic work and prepare for professional practice in music fields.",
            "keywords": ["research", "writing", "paper", "thesis", "seminar", "presentation", "professional", "career", "practice", "pedagogy", "teaching", "project", "portfolio"],
            "prefix_boost": {"MUS": 1.2},
        },
    ],
    "School of Public Policy": [
        {
            "competency": "Public policy analysis and governance",
            "definition": "Ability to analyze public policy, governance structures, institutions and policy-making processes.",
            "keywords": ["public policy", "policy", "governance", "government", "institution", "institutions", "public administration", "administration", "regulation", "reform", "decision"],
            "prefix_boost": {"PUB": 1.8, "URB": 0.7},
        },
        {
            "competency": "Urban management, planning and sustainable development",
            "definition": "Ability to understand urban management, planning, land use, infrastructure, sustainability and urban development.",
            "keywords": ["urban", "city", "cities", "planning", "land use", "infrastructure", "transport", "transportation", "housing", "sustainable", "sustainability", "development", "environment"],
            "prefix_boost": {"URB": 2.0, "PUB": 0.4},
        },
        {
            "competency": "Policy research methods, data and evaluation",
            "definition": "Ability to use qualitative, quantitative and data-oriented methods for policy research, programme evaluation and evidence-based decisions.",
            "keywords": ["research", "method", "methods", "data", "statistics", "statistical", "quantitative", "qualitative", "survey", "evaluation", "evidence", "empirical", "analysis", "model"],
            "prefix_boost": {"PUB": 1.2, "URB": 1.1},
        },
        {
            "competency": "Economics, finance and resource allocation for public problems",
            "definition": "Ability to reason about economics, public finance, budgeting, cost-benefit analysis and resource allocation in policy settings.",
            "keywords": ["economics", "economic", "finance", "financial", "budget", "budgeting", "cost", "benefit", "cost-benefit", "resource", "allocation", "market", "tax", "fiscal"],
            "prefix_boost": {"PUB": 0.9, "URB": 0.7},
        },
        {
            "competency": "Social, environmental and regional policy problem-solving",
            "definition": "Ability to address social, environmental, regional and community-level public problems through policy and governance tools.",
            "keywords": ["social", "society", "community", "environmental", "environment", "regional", "inequality", "poverty", "health", "education", "climate", "public problem"],
            "prefix_boost": {"PUB": 1.0, "URB": 0.9},
        },
        {
            "competency": "Leadership, communication and stakeholder engagement",
            "definition": "Ability to communicate policy ideas, lead teams, negotiate with stakeholders and present governance solutions.",
            "keywords": ["leadership", "communication", "presentation", "stakeholder", "negotiation", "collaboration", "team", "management", "public speaking", "writing", "report"],
            "prefix_boost": {"PUB": 0.8, "URB": 0.8},
        },
        {
            "competency": "Global, comparative and China policy perspectives",
            "definition": "Ability to compare policy systems and understand global, regional and China-related governance contexts.",
            "keywords": ["global", "international", "comparative", "china", "chinese", "regional", "cross-border", "governance system", "development model", "world"],
            "prefix_boost": {"PUB": 0.8, "URB": 0.7},
        },
    ],
    "School of Science and Engineering": [
        {
            "competency": "Mathematical, physical and quantitative science foundation",
            "definition": "Ability to use mathematics, physics, modelling and quantitative reasoning for science and engineering.",
            "keywords": ["mathematics", "math", "physics", "physical", "quantitative", "model", "modelling", "calculus", "linear algebra", "matrix", "stochastic", "probability"],
            "prefix_boost": {"CHM": 0.4, "CIE": 1.1, "CSC": 0.8},
        },
        {
            "competency": "General, analytical, inorganic, organic and physical chemistry",
            "definition": "Ability to understand core chemistry fields, including analytical, inorganic, organic and physical chemistry.",
            "keywords": ["chemistry", "chemical", "analytical", "inorganic", "organic", "physical chemistry", "spectroscopy", "chromatography", "reaction", "synthesis", "catalysis"],
            "prefix_boost": {"CHM": 2.4},
        },
        {
            "competency": "Chemistry laboratory, instrumentation and experimental practice",
            "definition": "Ability to perform laboratory work, instrumental analysis, experiments and scientific measurement.",
            "keywords": ["laboratory", "lab", "experiment", "experimental", "instrumental", "instrumentation", "analysis", "measurement", "technique", "techniques", "spectroscopic", "microscopy"],
            "prefix_boost": {"CHM": 1.6},
        },
        {
            "competency": "Materials science, polymers, functional materials and nanotechnology",
            "definition": "Ability to understand materials structure, functional materials, polymers, nanotechnology and material applications.",
            "keywords": ["materials", "material", "polymer", "functional materials", "nanotechnology", "nano", "structure", "properties", "supramolecular", "solid", "semiconductor"],
            "prefix_boost": {"CHM": 1.0, "CIE": 0.8},
        },
        {
            "competency": "Computing, programming, algorithms and data systems for engineering",
            "definition": "Ability to use programming, algorithms, computing systems and data tools in science and engineering.",
            "keywords": ["programming", "computer", "computing", "algorithm", "algorithms", "software", "data", "database", "system", "systems", "network", "cloud"],
            "prefix_boost": {"CSC": 2.0, "CIE": 1.5},
        },
        {
            "competency": "AI, machine learning, computer vision and data analytics",
            "definition": "Ability to use AI, machine learning, deep learning, computer vision and analytics in engineering contexts.",
            "keywords": ["artificial intelligence", "ai", "machine learning", "deep learning", "computer vision", "image", "data analytics", "data analysis", "neural", "classification", "prediction"],
            "prefix_boost": {"CIE": 1.7, "CSC": 1.5},
        },
        {
            "competency": "Electronics, circuits, communications and information systems",
            "definition": "Ability to understand circuits, electronics, communications, RF, IC design and information systems.",
            "keywords": ["electronics", "circuit", "circuits", "communication", "wireless", "rf", "cmos", "integrated circuit", "ic", "signal", "information theory", "network"],
            "prefix_boost": {"CIE": 2.1, "CSC": 0.4},
        },
        {
            "competency": "Control, robotics, intelligent systems and automation",
            "definition": "Ability to understand control theory, robotics, intelligent systems and automation.",
            "keywords": ["control", "robotics", "robot", "intelligent system", "intelligent systems", "automation", "kinematics", "sensor", "nanorobotics", "machine intelligence"],
            "prefix_boost": {"CIE": 1.8, "CSC": 0.6},
        },
        {
            "competency": "Energy systems, photonics and advanced engineering applications",
            "definition": "Ability to apply science and engineering to energy systems, photonics, devices and advanced technologies.",
            "keywords": ["energy", "photonics", "device", "devices", "engineering", "semiconductor", "advanced", "application", "applications", "technology"],
            "prefix_boost": {"CIE": 1.0, "CHM": 0.5},
        },
        {
            "competency": "Research, capstone and scientific project execution",
            "definition": "Ability to conduct research, projects, thesis work, enterprise practice and scientific reporting.",
            "keywords": ["research", "project", "capstone", "thesis", "design", "practice", "enterprise", "seminar", "presentation", "report"],
            "prefix_boost": {"CHM": 0.6, "CIE": 0.8, "CSC": 0.6},
        },
    ],
}


def normalize_code(code: str) -> str:
    return re.sub(r"\s+", "", code or "").upper()


def extract_ucore_codes() -> dict:
    by_pdf = {}
    all_codes = set()
    pattern = re.compile(r"\b[A-Z]{2,4}\s?\d{4}[A-Z]?\b")
    for pdf in CURRICULUM_PDFS:
        reader = PdfReader(str(pdf))
        text = "\n".join((page.extract_text() or "") for page in reader.pages)
        codes = {normalize_code(c) for c in pattern.findall(text)}
        by_pdf[pdf.name] = sorted(codes)
        all_codes.update(codes)
    return {"all_codes": sorted(all_codes), "by_pdf": by_pdf}


def read_competency_programmes() -> list[dict]:
    wb = load_workbook(COMPETENCY_XLSX, read_only=True, data_only=True)
    ws = wb["Competency Review"]
    headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    idx = {h: i for i, h in enumerate(headers)}
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[idx["Programme"]]:
            continue
        records.append({h: row[i] if i < len(row) else None for h, i in idx.items()})
    return records


def programme_applies_to_school(programme: dict, school_en: str) -> bool:
    if programme.get("School / Faculty") == school_en:
        return True
    note = f"{programme.get('School Notes') or ''} {programme.get('School / Faculty') or ''}"
    abbr = SCHOOL_ABBR.get(school_en)
    if abbr and re.search(rf"\b{re.escape(abbr)}\b", note):
        return True
    if programme.get("Programme") == "Financial Engineering" and school_en in {
        "School of Management and Economics", "School of Data Science", "School of Science and Engineering"
    }:
        return True
    return False


def build_pools(programmes: list[dict]) -> list[dict]:
    pools = []
    for school, items in POOL_DEFINITIONS.items():
        source_programmes = [
            p["Programme"] for p in programmes if programme_applies_to_school(p, school)
        ]
        for item in items:
            pools.append({
                "school": school,
                "school_abbr": SCHOOL_ABBR.get(school, ""),
                "competency": item["competency"],
                "definition": item["definition"],
                "keywords": item["keywords"],
                "source_programmes": source_programmes,
                "source_programmes_text": "; ".join(source_programmes),
                "prefix_boost": item.get("prefix_boost", {}),
            })
    return pools


def read_courses(ucore_codes: set[str]) -> tuple[list[dict], list[dict], list[dict]]:
    wb = load_workbook(COURSE_XLSX, read_only=True, data_only=True)
    ws = wb["With Outline"]
    headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    idx = {h: i for i, h in enumerate(headers)}
    courses, excluded, skipped_no_syllabus = [], [], []
    text_fields = [
        "course_title", "detail_description", "description_english", "description_chinese",
        "prerequisites", "co_requisites", "learning_outcomes", "course_syllabus",
        "assessment_scheme", "course_components", "offered_terms",
    ]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[idx["letter"]] != TARGET_LETTER:
            continue
        record = {h: row[i] for h, i in idx.items()}
        record["course_code"] = normalize_code(record.get("course_code"))
        record["school_en"] = SCHOOL_MAP.get(record.get("school") or "", record.get("school") or "")
        record["text_for_scoring"] = "\n".join(str(record.get(f) or "") for f in text_fields)
        if not str(record.get("course_syllabus") or "").strip():
            record["exclude_reason"] = "No course_syllabus field; only courses with syllabus are analyzed"
            skipped_no_syllabus.append(record)
            continue
        prefix_reason = next(
            (reason for prefix, reason in PREFIX_EXCLUSION_RULES.items() if str(record["course_code"]).startswith(prefix)),
            None,
        )
        if prefix_reason:
            record["exclude_reason"] = prefix_reason
            excluded.append(record)
            continue
        if record["course_code"] in ucore_codes:
            record["exclude_reason"] = "UCore course listed in supplied Academic Curriculum PDF(s)"
            excluded.append(record)
        else:
            courses.append(record)
    return courses, excluded, skipped_no_syllabus


def phrase_count(text: str, phrase: str) -> int:
    phrase_l = phrase.lower()
    text_l = text.lower()
    if " " in phrase_l or "-" in phrase_l:
        return text_l.count(phrase_l)
    return len(re.findall(rf"\b{re.escape(phrase_l)}s?\b", text_l))


def score_course_competency(course: dict, pool: dict) -> dict:
    text = (course.get("text_for_scoring") or "").lower()
    title = str(course.get("course_title") or "").lower()
    prefix = re.match(r"[A-Z]+", course["course_code"]).group(0)
    hits = []
    weighted_hit_score = 0.0
    for kw in pool["keywords"]:
        n = phrase_count(text, kw)
        if n:
            hits.append(kw)
            weighted_hit_score += min(3, n) * (1.35 if " " in kw else 1.0)
    coverage = len(set(hits)) / max(len(pool["keywords"]), 1)
    keyword_component = min(6.8, 2.5 * weighted_hit_score ** 0.55 + 2.0 * coverage)
    prefix_component = pool.get("prefix_boost", {}).get(prefix, 0.0)
    title_component = 0.0
    for kw in pool["keywords"]:
        if phrase_count(title, kw):
            title_component = min(1.2, title_component + 0.45)
    score = max(0.0, min(10.0, keyword_component + prefix_component + title_component))
    if score >= 8:
        level = "Very high"
    elif score >= 6:
        level = "High"
    elif score >= 4:
        level = "Moderate"
    elif score >= 2:
        level = "Low"
    else:
        level = "Very low"
    rationale = []
    if hits:
        rationale.append("course outline matches " + ", ".join(hits[:10]))
    if prefix_component:
        rationale.append(f"{prefix} prefix aligns with this competency area (+{prefix_component:.1f})")
    if title_component:
        rationale.append("course title directly signals this area")
    if not rationale:
        rationale.append("limited explicit evidence in the course outline")
    return {
        "score": round(score, 1),
        "level": level,
        "matched_keywords": ", ".join(sorted(set(hits))),
        "keyword_component": round(keyword_component, 2),
        "prefix_component": round(prefix_component, 2),
        "title_component": round(title_component, 2),
        "rationale": "; ".join(rationale),
    }


def main() -> None:
    ucore = extract_ucore_codes()
    programmes = read_competency_programmes()
    pools = build_pools(programmes)
    courses, excluded, skipped_no_syllabus = read_courses(set(ucore["all_codes"]))
    rows = []
    for course in courses:
        course_pools = [p for p in pools if p["school"] == course["school_en"]]
        for pool in course_pools:
            scored = score_course_competency(course, pool)
            rows.append({
                "course_code": course["course_code"],
                "course_title": course.get("course_title") or "",
                "subject": course.get("subject") or "",
                "offering_school": course.get("school") or "",
                "offering_school_en": course.get("school_en") or "",
                "academic_org": course.get("academic_org") or "",
                "competency_pool_school": pool["school"],
                "competency": pool["competency"],
                "competency_definition": pool["definition"],
                "source_programmes": pool["source_programmes_text"],
                **scored,
                "course_description": course.get("description_english") or course.get("detail_description") or "",
                "learning_outcomes": course.get("learning_outcomes") or "",
                "course_syllabus": course.get("course_syllabus") or "",
            })
    rows.sort(key=lambda r: (r["course_code"], -r["score"], r["competency"]))
    OUT_JSON.write_text(json.dumps({
        "method": {
            "scope": f"{TARGET_LETTER}-letter SIS courses from sis_course_outlines_export.xlsx with non-empty course_syllabus; UCore courses in supplied Academic Curriculum PDFs and all CEC/GEB/GEC/GED/GEW-prefix courses are excluded.",
            "unit_of_analysis": "course-to-faculty-competency-pool item, not course-to-programme",
            "score_scale": "0-10",
            "score_components": ["course outline keyword/phrase evidence", "course prefix alignment", "course title signal"],
        },
        "ucore_codes": ucore,
        "competency_pools": [
            {k: v for k, v in p.items() if k not in {"prefix_boost"}}
            for p in pools
        ],
        "scored_rows": rows,
        "excluded_ucore_courses": excluded,
        "skipped_no_syllabus_courses": skipped_no_syllabus,
        "summary": {
            "courses_scored": len({r["course_code"] for r in rows}),
            "score_rows": len(rows),
            "competency_pool_items": len(pools),
            "ucore_courses_excluded": len(excluded),
            "courses_skipped_no_syllabus": len(skipped_no_syllabus),
        },
    }, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"out": str(OUT_JSON), **json.loads(OUT_JSON.read_text(encoding='utf-8'))["summary"]}, ensure_ascii=False))


if __name__ == "__main__":
    main()
